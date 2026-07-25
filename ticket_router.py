import os
import re
import csv
import openpyxl
from openpyxl.utils import get_column_letter

# 1. The Regular Expression verification rule
TICKET_PATTERN = r"^(ERROR|DEBUG|INFO|WARNING)\s+\[pid:(\d+)\]\s+(.*?)\s+-\s+User:(EMP\d{4})$"


# 2. NEW MODIFICATION: The auto-fitting Excel layout manager
def convert_csv_to_beautiful_excel(csv_path, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IT Dashboard"

    # Ingest the CSV table data
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Calculate text length and automatically expand columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        # Add comfort padding to prevent layout clipping
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(excel_path)
    print(f"✨ Auto-scaled Excel dashboard generated at: {excel_path}")


# 3. Your Unified Core Router Function
def generate_unified_dashboard(input_dir, output_csv, output_excel):
    if not os.path.exists(input_dir):
        print(f"❌ Error: The input directory '{input_dir}' does not exist.")
        return

    csv_headers = ["Ticket ID", "Priority", "System Message", "Assigned Employee"]
    total_lines = 0
    clean_lines = 0
    malformed_lines = 0

    print("📊 COMPILING UNIFIED STRUCTURED IT DASHBOARD...")

    with open(output_csv, "w", newline="", encoding="utf-8") as csv_out:
        writer = csv.writer(csv_out)
        writer.writerow(csv_headers)

        for filename in os.listdir(input_dir):
            if filename.endswith(".log"):
                file_path = os.path.join(input_dir, filename)

                with open(file_path, "r", encoding="utf-8") as log_file:
                    for line_num, line in enumerate(log_file, 1):
                        clean_line = line.strip()
                        if not clean_line:
                            continue

                        total_lines += 1
                        match = re.match(TICKET_PATTERN, clean_line)

                        if match:
                            clean_lines += 1
                            writer.writerow([
                                f"TICK-{match.group(2)}",
                                match.group(1),
                                match.group(3),
                                match.group(4)
                            ])
                        else:
                            malformed_lines += 1
                            writer.writerow([
                                "MALFORMED-LOG",
                                "AUDIT_REQUIRED",
                                f"[Line {line_num}] Raw Data: {clean_line}",
                                "SYSTEM_ADMIN"
                            ])

    print("=" * 60)
    print(f"🏁 CSV Generation Complete: {output_csv}")
    print(f"  ↳ Total Records Evaluated: {total_lines}")
    print(f"  ↳ Clean Parsed Logs:       {clean_lines}")
    print(f"  ↳ Malformed Logs Flagged:   {malformed_lines}")
    print("=" * 60)

    # Automatically switch on the Excel modification right here
    convert_csv_to_beautiful_excel(output_csv, output_excel)


# 4. Global Target Configuration Block
if __name__ == "__main__":
    INPUT_FOLDER = "unprocessed_tickets"
    OUTPUT_CSV = "structured_it_dashboard.csv"
    OUTPUT_EXCEL = "structured_it_dashboard.xlsx"

    generate_unified_dashboard(INPUT_FOLDER, OUTPUT_CSV, OUTPUT_EXCEL)